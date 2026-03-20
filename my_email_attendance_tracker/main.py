import openpyxl
import smtplib
wb=openpyxl.load_workbook("sample_workbook.xlsx")
sheet=wb.active
# for row in sheet.iter_rows(min_row=1, max_row=4, values_only=True):
#     print(row)

cellvalue=0
n=2
while n<5:
    studentemail=sheet.cell(row=n,column=2).value
    print(studentemail)
    cellvalue=sheet.cell(row=n,column=3).value
    print(cellvalue)
    if int(cellvalue)>1:
        with smtplib.SMTP("smtp.gmail.com") as email:
            email.starttls()
            email.login(user="rudrakshswami931@gmail.com", password="passward")
            email.sendmail(
                from_addr="rudrakshswami931@gmail.com",
                to_addrs=f"{studentemail}",
                msg="Subject:last reminder\n\nyou have taken 2 leaves this mouths in python lec if you get aone more leave you will be officially can not give mse exam, this is your last reminder."
            )
    n+=1




cellvalu=0
m=2
while m<5:
    studentemail=sheet.cell(row=m,column=2).value
    print(studentemail)
    cellvalu=sheet.cell(row=m,column=4).value
    print(cellvalu)
    if int(cellvalu)>=2:
        with smtplib.SMTP("smtp.gmail.com") as email:
            email.starttls()
            email.login(user="rudrakshswami931@gmail.com", password="passward")
            email.sendmail(
                from_addr="rudrakshswami931@gmail.com",
                to_addrs=f"{studentemail}",
                msg="Subject:last reminder\n\nyou have taken 2 leaves this mouths in maths lec if you get aone more leave you will be officially can not give mse exam, this is your last reminder."
            )
    m+=1



cellval=0
o=2
while o<5:
    studentemail=sheet.cell(row=o,column=2).value
    print(studentemail)
    cellval=sheet.cell(row=o,column=5).value
    print(cellval)
    if int(cellval)>=2:
        with smtplib.SMTP("smtp.gmail.com") as email:
            email.starttls()
            email.login(user="rudrakshswami931@gmail.com", password="passward")
            email.sendmail(
                from_addr="rudrakshswami931@gmail.com",
                to_addrs=f"{studentemail}",
                msg="Subject:last reminder\n\nyou have taken 2 leaves this mouths in chem lec if you get aone more leave you will be officially can not give mse exam, this is your last reminder."
            )
    o+=1



















# wb.save("sample_workbook.xlsx")
# with smtplib.SMTP("smtp.gmail.com") as email:
#     email.starttls()
#     email.login(user="rudrakshswami931@gmail.com",password="znoa qoyp uzoh lshk")
#     email.sendmail(
#         from_addr="rudrakshswami931@gmail.com",
#         to_addrs="rudrakshdswami@gmail.com",
#         msg="Subject:about your attendance\n\nyou have taken 2 leaves this mouths if you get aone more leave you will be officially can not give mse exam, this is your last reminder."
#     )

